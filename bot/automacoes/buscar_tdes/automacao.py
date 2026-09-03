from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from servicos.fraction import (
    login_fraction,
    pesquisar_e_capturar_tdes_fraction,
)


PASTA_FLUXO = Path(__file__).resolve().parent
PASTA_BOT = Path(__file__).resolve().parents[2]
PASTA_RAIZ = PASTA_BOT.parent

PASTA_RESULTADOS = (
    PASTA_RAIZ
    / "resultados"
    / "tdes"
)

load_dotenv(
    PASTA_BOT / ".env",
    override=False,
)

FRACTION_USER = os.getenv(
    "FRACTION_USER",
    "",
).strip()

FRACTION_PASSWORD = os.getenv(
    "FRACTION_PASSWORD",
    "",
).strip()

HEADLESS = (
    os.getenv("HEADLESS", "True")
    .strip()
    .lower()
    in {"1", "true", "yes", "sim", "on"}
)

PESO_MINIMO = 30.0


def validar_env() -> None:
    faltantes = []

    if not FRACTION_USER:
        faltantes.append("FRACTION_USER")

    if not FRACTION_PASSWORD:
        faltantes.append("FRACTION_PASSWORD")

    if faltantes:
        raise RuntimeError(
            "Variável(is) ausente(s) no .env: "
            + ", ".join(faltantes)
        )


def normalizar_cte(valor: object) -> str:
    if pd.isna(valor):
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    texto = str(valor).strip()

    if re.fullmatch(r"\d+\.0", texto):
        return texto[:-2]

    return texto


def normalizar_peso(valor: object) -> float | None:
    if pd.isna(valor):
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = (
        str(valor)
        .strip()
        .lower()
        .replace("kg", "")
        .replace(" ", "")
    )

    if not texto:
        return None

    # Formato brasileiro:
    # 40,5
    # 42,24
    # Também aceita 30 ou 30.5.
    if "," in texto:
        texto = texto.replace(".", "")
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return None


def preparar_base(
    arquivo,
) -> pd.DataFrame:
    arquivo.seek(0)

    df = pd.read_excel(
        arquivo,
        dtype=object,
    )

    arquivo.seek(0)

    if len(df.columns) < 2:
        raise ValueError(
            "A planilha precisa possuir as colunas CTE e PESO."
        )

    primeira = str(df.columns[0]).strip().upper()
    segunda = str(df.columns[1]).strip().upper()

    if primeira != "CTE" or segunda != "PESO":
        raise ValueError(
            "O arquivo precisa possuir A1 = CTE e B1 = PESO."
        )

    coluna_cte = df.columns[0]
    coluna_peso = df.columns[1]

    preparada = pd.DataFrame(
        {
            "CTE": df[coluna_cte].map(normalizar_cte),
            "Peso": df[coluna_peso].map(normalizar_peso),
        }
    )

    preparada = preparada[
        (preparada["CTE"] != "")
        & preparada["Peso"].notna()
        & (preparada["Peso"] >= PESO_MINIMO)
    ].copy()

    preparada.reset_index(
        drop=True,
        inplace=True,
    )

    if preparada.empty:
        raise ValueError(
            "Nenhum pedido com peso maior ou igual a 30 kg foi encontrado."
        )

    return preparada


def formatar_descricoes(
    descricoes: list[str],
) -> str:
    descricoes_limpas = [
        " ".join(str(descricao).split())
        for descricao in descricoes
        if str(descricao).strip()
    ]

    return "\n\n".join(
        f"{indice}° {descricao}"
        for indice, descricao in enumerate(
            descricoes_limpas,
            start=1,
        )
    )


def salvar_resultado(
    df: pd.DataFrame,
) -> Path:
    PASTA_RESULTADOS.mkdir(
        parents=True,
        exist_ok=True,
    )

    nome = (
        "tdes_"
        + datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )
        + ".xlsx"
    )

    caminho = (
        PASTA_RESULTADOS
        / nome
    )

    df.to_excel(
        caminho,
        index=False,
    )

    wb = load_workbook(caminho)
    ws = wb.active

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:C{ws.max_row}"
    )

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80

    for celula in ws[1]:
        celula.font = Font(
            bold=True,
        )
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for linha in range(
        2,
        ws.max_row + 1,
    ):
        ws.cell(
            linha,
            1,
        ).number_format = "@"

        ws.cell(
            linha,
            2,
        ).number_format = '0.00'

        descricao = ws.cell(
            linha,
            3,
        )

        descricao.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        quantidade_linhas = max(
            1,
            str(descricao.value or "").count("\n") + 1,
        )

        ws.row_dimensions[
            linha
        ].height = max(
            20,
            18 * quantidade_linhas,
        )

    wb.save(caminho)

    return caminho


def executar_busca_tdes(
    arquivo,
    log: Callable[[str], None] = print,
) -> dict:
    validar_env()

    base = preparar_base(
        arquivo
    )

    total = len(base)

    log(
        f"{total} pedido(s) com peso >= 30 kg."
    )

    cache_descricoes: dict[str, str] = {}

    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=HEADLESS,
            slow_mo=0 if HEADLESS else 300,
        )

        context = browser.new_context(
            no_viewport=not HEADLESS,
        )

        page = context.new_page()

        try:
            login_fraction(
                page,
                FRACTION_USER,
                FRACTION_PASSWORD,
                log,
            )

            for posicao, linha in base.iterrows():
                cte = linha["CTE"]

                log(
                    f"[{posicao + 1}/{total}] CTE {cte}"
                )

                if cte in cache_descricoes:
                    log(
                        f"CTE {cte}: resultado reutilizado."
                    )
                    continue

                try:
                    descricoes = (
                        pesquisar_e_capturar_tdes_fraction(
                            page,
                            cte,
                            log,
                        )
                    )

                    cache_descricoes[cte] = (
                        formatar_descricoes(
                            descricoes
                        )
                    )

                except Exception as erro:
                    cache_descricoes[cte] = ""

                    log(
                        f"CTE {cte}: erro na consulta: "
                        f"{type(erro).__name__}: {erro}"
                    )

        finally:
            context.close()
            browser.close()

    resultado = base.copy()

    resultado["Descrição TDE"] = (
        resultado["CTE"]
        .map(cache_descricoes)
        .fillna("")
    )

    resultado = resultado[
        [
            "CTE",
            "Peso",
            "Descrição TDE",
        ]
    ]

    caminho = salvar_resultado(
        resultado
    )

    log(
        f"Resultado salvo em: {caminho}"
    )

    return {
        "arquivo": caminho,
        "total": total,
        "resultado": resultado,
    }
