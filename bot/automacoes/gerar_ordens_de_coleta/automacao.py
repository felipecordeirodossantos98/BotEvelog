from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import (
    Locator,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


PASTA_FLUXO = Path(__file__).resolve().parent

# bot/
PASTA_BOT = Path(__file__).resolve().parents[2]

# BotEvelog/
PASTA_RAIZ = PASTA_BOT.parent


PASTA_DADOS = (
    PASTA_FLUXO
    / "dados"
)

PASTA_RESULTADOS = (
    PASTA_RAIZ
    / "resultados"
    / "ordens_de_coleta"
)

PASTA_PERFIS = (
    PASTA_RAIZ
    / "perfis"
)

PASTA_PERFIL = (
    PASTA_PERFIS
    / "gerar_ordens_de_coleta"
    / "chromium"
)


ARQUIVO_ENV = (
    PASTA_BOT
    / ".env"
)

ARQUIVO_BASE_CNPJS = (
    PASTA_DADOS
    / "base_cnpjs.json"
)

ARQUIVO_EMAILS = (
    PASTA_DADOS
    / "emails_unidades.json"
)


load_dotenv(
    ARQUIVO_ENV,
    override=False,
)

URL = os.getenv(
    "URL_FRACTION",
    "",
).strip()

TIMEOUT = 30_000

MAX_TENTATIVAS = 3

CNPJ_DESTINATARIO = "42591651000143"

# Valores definidos durante os testes do formulário.
CONTA_CORRENTE = "0153080"
OBSERVACAO = "bag"
CONTEUDO = "bag - malote"
VOLUMES = "1"
PESO = "1,00"
VALOR_COLETA = "13,20"
NUMERO_NOTA = "dec"
SERIE = "0"
VALOR_NOTA = "100,00"
MODALIDADE = "CORPORATE"

def carregar_emails_unidades() -> dict[str, str]:
    if not ARQUIVO_EMAILS.exists():
        raise FileNotFoundError(
            f"Base de e-mails não encontrada: {ARQUIVO_EMAILS}"
        )

    with ARQUIVO_EMAILS.open(
        "r",
        encoding="utf-8",
    ) as arquivo:
        dados = json.load(arquivo)

    if not isinstance(dados, dict):
        raise ValueError(
            "emails_unidades.json deve possuir o formato "
            '{"UNIDADE": ["email@dominio.com.br"]}.'
        )

    emails: dict[str, str] = {}

    for unidade, valores in dados.items():
        unidade = texto_limpo(unidade)

        if not unidade:
            continue

        if isinstance(valores, list):
            lista_emails = [
                texto_limpo(email)
                for email in valores
                if texto_limpo(email)
            ]
        elif isinstance(valores, str):
            lista_emails = [
                texto_limpo(valores)
            ] if texto_limpo(valores) else []
        else:
            raise ValueError(
                f"Formato de e-mail inválido para a unidade {unidade}."
            )

        emails[unidade] = "; ".join(lista_emails)

    return emails

def criar_planilha_erros(
    detalhes: list[dict],
) -> Path | None:
    erros = [
        item
        for item in detalhes
        if item["SITUAÇÃO"] == "FALHA"
    ]

    if not erros:
        return None

    df_erros = pd.DataFrame(erros)

    resumo = (
        df_erros
        .groupby(
            "SIGLA",
            as_index=False,
        )
        .size()
        .rename(
            columns={
                "SIGLA": "SIGLA DO RESTAURANTE",
                "size": "NUMERO DE ORDENS",
            }
        )
    )

    agora = datetime.now()

    nome_arquivo = agora.strftime(
        "ordens_erro_%d-%m-%Y_%H-%M-%S.xlsx"
    )

    caminho = PASTA_RESULTADOS / nome_arquivo

    wb = Workbook()
    ws = wb.active
    ws.title = "Ordens com Erro"

    ws.append(
        [
            "SIGLA DO RESTAURANTE",
            "NUMERO DE ORDENS",
        ]
    )

    for _, linha in resumo.iterrows():
        ws.append(
            [
                linha["SIGLA DO RESTAURANTE"],
                int(linha["NUMERO DE ORDENS"]),
            ]
        )

    azul_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="8EDDE1",
    )

    vermelho_erro = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    fonte_erro = Font(
        color="9C0006",
    )

    borda_fina = Border(
        left=Side(
            style="thin",
            color="000000",
        ),
        right=Side(
            style="thin",
            color="000000",
        ),
        top=Side(
            style="thin",
            color="000000",
        ),
        bottom=Side(
            style="thin",
            color="000000",
        ),
    )

    for celula in ws[1]:
        celula.fill = azul_cabecalho
        celula.font = Font(
            bold=True,
        )
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        celula.border = borda_fina

    for linha in range(
        2,
        ws.max_row + 1,
    ):
        for coluna in range(
            1,
            3,
        ):
            celula = ws.cell(
                linha,
                coluna,
            )

            celula.fill = vermelho_erro
            celula.font = fonte_erro
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            celula.border = borda_fina

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:B{ws.max_row}"
    )

    wb.save(caminho)

    return caminho


def normalizar_coluna(valor: object) -> str:
    return " ".join(str(valor).strip().upper().split())


def normalizar_sigla(valor: object) -> str:
    return " ".join(str(valor).strip().upper().split())


def texto_limpo(valor: object) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip()

    if texto.lower() == "nan":
        return ""

    return texto


def somente_digitos(valor: object) -> str:
    return "".join(
        caractere
        for caractere in texto_limpo(valor)
        if caractere.isdigit()
    )


def ler_excel_normalizado(caminho_ou_arquivo) -> pd.DataFrame:
    df = pd.read_excel(
        caminho_ou_arquivo,
        dtype=str,
    ).fillna("")

    df.columns = [
        normalizar_coluna(coluna)
        for coluna in df.columns
    ]

    return df


def validar_arquivos_fixos() -> list[str]:
    erros: list[str] = []

    if not ARQUIVO_ENV.exists():
        erros.append(
            "Crie o arquivo .env na raiz do projeto."
        )
    else:
        usuario = os.getenv("FRACTION_USER", "").strip()
        senha = os.getenv("FRACTION_PASSWORD", "").strip()
        url_fraction = os.getenv("URL_FRACTION", "").strip()

        if not usuario:
            erros.append(
                "FRACTION_USER não foi informado no .env."
            )

        if not senha:
            erros.append(
                "FRACTION_PASSWORD não foi informado no .env."
            )

        if not url_fraction:
            erros.append(
                "URL_FRACTION não foi informada no .env."
            )

    if not ARQUIVO_BASE_CNPJS.exists():
        erros.append(
            "Crie dados/base_cnpjs.json."
        )
    else:
        try:
            carregar_base_cnpjs()
        except Exception as erro:
            erros.append(
                f"Não foi possível ler base_cnpjs.json: {erro}"
            )

    if not ARQUIVO_EMAILS.exists():
        erros.append(
            "Crie dados/emails_unidades.json."
        )
    else:
        try:
            carregar_emails_unidades()
        except Exception as erro:
            erros.append(
                f"Não foi possível ler emails_unidades.json: {erro}"
            )

    return erros

def validar_planilha_pedidos(df_original: pd.DataFrame) -> list[str]:
    erros: list[str] = []

    df = df_original.copy()
    df.columns = [
        normalizar_coluna(coluna)
        for coluna in df.columns
    ]

    obrigatorias = {
        "SIGLA DO RESTAURANTE",
        "NUMERO DE ORDENS",
    }

    faltantes = obrigatorias - set(df.columns)

    if faltantes:
        erros.append(
            "Colunas ausentes: "
            + ", ".join(sorted(faltantes))
        )
        return erros

    if df.empty:
        erros.append("A base de pedidos está vazia.")
        return erros

    for indice, linha in df.iterrows():
        numero_linha = indice + 2
        sigla = normalizar_sigla(
            linha["SIGLA DO RESTAURANTE"]
        )
        quantidade_texto = texto_limpo(
            linha["NUMERO DE ORDENS"]
        )

        if not sigla:
            erros.append(
                f"Linha {numero_linha}: sigla vazia."
            )

        try:
            quantidade = int(float(quantidade_texto))
        except (TypeError, ValueError):
            erros.append(
                f"Linha {numero_linha}: NUMERO DE ORDENS inválido."
            )
            continue

        if quantidade <= 0:
            erros.append(
                f"Linha {numero_linha}: NUMERO DE ORDENS "
                "deve ser maior que zero."
            )

    return erros


def carregar_login() -> tuple[str, str]:
    usuario = os.getenv(
        "FRACTION_USER",
        "",
    ).strip()

    senha = os.getenv(
        "FRACTION_PASSWORD",
        "",
    ).strip()

    return usuario, senha


def carregar_base_cnpjs() -> dict[str, str]:
    if not ARQUIVO_BASE_CNPJS.exists():
        raise FileNotFoundError(
            f"Arquivo não encontrado: {ARQUIVO_BASE_CNPJS}"
        )

    with ARQUIVO_BASE_CNPJS.open(
        "r",
        encoding="utf-8",
    ) as arquivo:
        dados = json.load(arquivo)

    if not isinstance(dados, dict):
        raise ValueError(
            "base_cnpjs.json deve possuir o formato "
            '{"SIGLA": "CNPJ"}.'
        )

    base: dict[str, str] = {}

    for sigla, cnpj in dados.items():
        sigla = normalizar_sigla(sigla)
        cnpj = somente_digitos(cnpj)

        if not sigla:
            continue

        if len(cnpj) != 14:
            raise ValueError(
                f"CNPJ inválido para a sigla {sigla}: {cnpj}"
            )

        if sigla in base:
            raise ValueError(
                f"A sigla {sigla} aparece mais de uma vez "
                "em base_cnpjs.json."
            )

        base[sigla] = cnpj

    return base

def preparar_execucoes(
    df_original: pd.DataFrame,
) -> tuple[list[dict], list[str]]:
    df = df_original.copy()
    df.columns = [
        normalizar_coluna(coluna)
        for coluna in df.columns
    ]

    base_cnpjs = carregar_base_cnpjs()

    execucoes: list[dict] = []
    alertas: list[str] = []

    for indice, linha in df.iterrows():
        sigla = normalizar_sigla(
            linha["SIGLA DO RESTAURANTE"]
        )
        quantidade = int(
            float(texto_limpo(linha["NUMERO DE ORDENS"]))
        )

        cnpj = base_cnpjs.get(sigla)

        if cnpj is None:
            alertas.append(
                f"Linha {indice + 2}: sigla {sigla} não encontrada "
                "em base_cnpjs.json."
            )
            continue

        for sequencia in range(1, quantidade + 1):
            execucoes.append(
                {
                    "sigla": sigla,
                    "sequencia": sequencia,
                    "total_sigla": quantidade,
                    "cnpj_remetente": cnpj,
                }
            )

    return execucoes, alertas


def digitar(campo: Locator, valor: str) -> None:
    campo.wait_for(
        state="visible",
        timeout=TIMEOUT,
    )
    campo.scroll_into_view_if_needed()
    campo.click()
    campo.press("Control+A")
    campo.press("Backspace")
    campo.type(
        texto_limpo(valor),
        delay=50,
    )


def colar_sem_tab(
    page: Page,
    campo: Locator,
    valor: str,
) -> None:
    campo.wait_for(
        state="visible",
        timeout=TIMEOUT,
    )
    campo.scroll_into_view_if_needed()
    campo.click()
    campo.press("Control+A")
    campo.press("Backspace")

    page.evaluate(
        "valor => navigator.clipboard.writeText(valor)",
        texto_limpo(valor),
    )

    campo.press("Control+V")
    page.wait_for_timeout(400)


def realizar_login(
    page: Page,
    usuario: str,
    senha: str,
    log: Callable[[str], None],
) -> None:
    log("Abrindo Jadlog...")

    page.goto(
        URL,
        wait_until="domcontentloaded",
        timeout=60_000,
    )

    try:
        campo_usuario = page.get_by_role(
            "textbox",
            name="Usuário",
            exact=True,
        )

        campo_usuario.wait_for(
            state="visible",
            timeout=7_000,
        )

        log("Realizando login...")

        campo_usuario.fill(usuario)

        page.get_by_role(
            "textbox",
            name="Senha",
            exact=True,
        ).fill(senha)

        page.get_by_role("button").click()

    except PlaywrightTimeoutError:
        log(
            "Tela de login não apareceu. "
            "A sessão pode já estar autenticada."
        )

    page.get_by_role(
        "link",
        name="Operacional",
    ).wait_for(
        state="visible",
        timeout=60_000,
    )

    log("Login concluído.")


def abrir_menu_com_submenu(
    menu: Locator,
    submenu: Locator,
    nome_menu: str,
) -> None:
    menu.wait_for(
        state="visible",
        timeout=TIMEOUT,
    )

    for tentativa in range(1, 4):
        menu.click()

        try:
            submenu.wait_for(
                state="visible",
                timeout=3_000,
            )
            return

        except PlaywrightTimeoutError:
            if tentativa == 3:
                raise RuntimeError(
                    f"Não foi possível abrir o menu {nome_menu}."
                )

def abrir_solicitacao_coleta(
    page: Page,
    log: Callable[[str], None],
) -> None:
    log("Abrindo Operacional...")

    menu_operacional = page.get_by_role(
        "link",
        name="Operacional",
    )

    submenu_ordem = page.get_by_role(
        "link",
        name="Ordem de coleta",
    )

    abrir_menu_com_submenu(
        menu_operacional,
        submenu_ordem,
        "Operacional",
    )

    log("Abrindo Ordem de coleta...")

    submenu_solicitacao = page.get_by_role(
        "link",
        name="Solicitação de Coleta",
    )

    abrir_menu_com_submenu(
        submenu_ordem,
        submenu_solicitacao,
        "Ordem de coleta",
    )

    log("Abrindo Solicitação de Coleta...")

    submenu_solicitacao.click()

    page.get_by_role(
        "textbox",
        name="Conta Corrente",
        exact=True,
    ).wait_for(
        state="visible",
        timeout=TIMEOUT,
    )

    log("Tela de Solicitação de Coleta aberta.")


def localizar_remetente(page: Page) -> Locator:
    return (
        page.get_by_role(
            "cell",
            name="Remetente CNPJ/CPF: Insc.",
        )
        .get_by_label("CNPJ/CPF:")
    )


def localizar_destinatario(page: Page) -> Locator:
    return (
        page.get_by_role(
            "cell",
            name="Destinatário CNPJ/CPF: Insc.",
        )
        .get_by_label("CNPJ/CPF:")
    )


def selecionar_modalidade_corporate(page: Page) -> None:
    modalidade = page.locator(
        '[id="form_emissao:modalidadeSelect"]'
    )

    modalidade.wait_for(
        state="visible",
        timeout=TIMEOUT,
    )

    caixa = modalidade.bounding_box()

    if caixa is None:
        raise RuntimeError(
            "Não foi possível localizar o campo Modalidade."
        )

    page.mouse.click(
        caixa["x"] + caixa["width"] - 10,
        caixa["y"] + caixa["height"] / 2,
    )

    page.wait_for_timeout(600)

    lista = page.locator(
        '[id="form_emissao:modalidadeSelect_items"]:visible'
    )

    lista.wait_for(
        state="visible",
        timeout=10_000,
    )

    lista.get_by_role(
        "option",
        name=MODALIDADE,
        exact=True,
    ).click()

    page.wait_for_timeout(700)


def preencher_formulario(
    page: Page,
    cnpj_remetente: str,
) -> None:
    digitar(
        page.get_by_role(
            "textbox",
            name="Conta Corrente",
            exact=True,
        ),
        CONTA_CORRENTE,
    )

    digitar(
        page.locator(
            '[id="form_emissao:observacaoArea"]'
        ),
        OBSERVACAO,
    )

    digitar(
        page.locator(
            '[id="form_emissao:conteudoArea"]'
        ),
        CONTEUDO,
    )

    # Seletores validados durante os testes.
    digitar(
        page.get_by_role(
            "textbox",
            name="ALT+7",
            exact=True,
        ),
        PESO,
    )

    digitar(
        page.get_by_role(
            "textbox",
            name="ALT+8",
            exact=True,
        ),
        VALOR_COLETA,
    )

    digitar(
        page.get_by_role(
            "textbox",
            name="ALT+9",
            exact=True,
        ),
        NUMERO_NOTA,
    )

    # Volume e série já aparecem com 1 e 0 na página.
    # Os valores são reforçados diretamente quando os IDs existem.
    campo_volume = page.locator(
        '[id="form_emissao:quantidadeVolume_input"]'
    )

    if campo_volume.count() > 0:
        digitar(campo_volume, VOLUMES)

    campo_serie = page.locator(
        '[id="form_emissao:bonfs:0:nota_serie"]'
    )

    if campo_serie.count() > 0:
        digitar(campo_serie, SERIE)

    digitar(
        page.locator(
            '[id="form_emissao:bonfs:0:nota_valor"]'
        ),
        VALOR_NOTA,
    )

    # Os dois CNPJs são colados antes de sair dos campos.
    colar_sem_tab(
        page,
        localizar_remetente(page),
        cnpj_remetente,
    )

    colar_sem_tab(
        page,
        localizar_destinatario(page),
        CNPJ_DESTINATARIO,
    )

    # Modalidade por último. O clique fora dispara o carregamento
    # dos dados de remetente e destinatário.
    selecionar_modalidade_corporate(page)

    page.wait_for_timeout(4_000)

    modalidade_exibida = page.locator(
        '[id="form_emissao:modalidadeSelect_label"]'
    ).inner_text()

    if modalidade_exibida.strip().upper() != MODALIDADE:
        raise RuntimeError(
            f"Modalidade inesperada: {modalidade_exibida}"
        )

def capturar_unidade_coletora(page: Page) -> str:
    texto_unidade = page.get_by_text(
        "Unidade coletora",
        exact=True,
    ).first

    texto_unidade.wait_for(
        state="visible",
        timeout=TIMEOUT,
    )

    celula = texto_unidade.locator(
        "xpath=ancestor::td[1]"
    )

    texto_completo = celula.inner_text().strip()

    linhas = [
        linha.strip()
        for linha in texto_completo.splitlines()
        if linha.strip()
        and linha.strip().lower() != "unidade coletora"
    ]

    if linhas:
        unidade = " ".join(linhas)

        return re.sub(
            r"\s+",
            " ",
            unidade,
        ).strip()

    # Alternativa para o caso de o nome estar
    # em uma célula anterior.
    celula_anterior = texto_unidade.locator(
        "xpath=ancestor::td[1]/preceding-sibling::td[1]"
    )

    if celula_anterior.count() > 0:
        unidade = celula_anterior.inner_text().strip()

        if unidade:
            return re.sub(
                r"\s+",
                " ",
                unidade,
            ).strip()

    raise RuntimeError(
        "A Unidade coletora apareceu, "
        "mas o nome da unidade não foi capturado."
    )

def gerar_coleta_e_capturar_ordem(page: Page) -> str:
    page.get_by_role(
        "button",
        name="Gerar Coleta",
        exact=True,
    ).click()

    mensagem = page.get_by_text(
        re.compile(
            r"Gerado\s+o\s+número\s+de\s+coleta",
            re.IGNORECASE,
        )
    ).last

    mensagem.wait_for(
        state="visible",
        timeout=30_000,
    )

    texto = mensagem.inner_text()

    resultado = re.search(
        r"Gerado\s+o\s+número\s+de\s+coleta\s+(\d+)",
        texto,
        flags=re.IGNORECASE,
    )

    if not resultado:
        # Fallback no contêiner branco mapeado pelo usuário.
        caixa = (
            page.locator("div")
            .filter(
                has_text=re.compile(
                    r"Gerado\s+o\s+número\s+de\s+coleta",
                    re.IGNORECASE,
                )
            )
            .nth(2)
        )

        caixa.wait_for(
            state="visible",
            timeout=10_000,
        )

        texto = caixa.inner_text()

        resultado = re.search(
            r"Gerado\s+o\s+número\s+de\s+coleta\s+(\d+)",
            texto,
            flags=re.IGNORECASE,
        )

    if not resultado:
        raise RuntimeError(
            "A coleta foi enviada, mas o número não pôde ser "
            f"extraído da mensagem: {texto}"
        )

    return resultado.group(1)


def criar_planilha_resultados(
    resultados: list[dict],
) -> Path:
    PASTA_RESULTADOS.mkdir(
        parents=True,
        exist_ok=True,
    )

    agora = datetime.now()
    nome = agora.strftime(
        "novas_ordens_%d-%m-%Y_%H-%M-%S.xlsx"
    )
    caminho = PASTA_RESULTADOS / nome

    wb = Workbook()
    ws = wb.active
    ws.title = "Novas Ordens"

    cabecalhos = [
        "RE",
        "SIGLA",
        "TIPO",
        "CTE",
        "VINCULAR/ ACERTO",
        "ORDEM",
        "SITUAÇÃO",
        "DT DE FINALIZAÇÃO",
        "DIAS FALTANTES",
        "SITUAÇÃO DE COLETA",
        "UNIDADE",
        "E-MAIL",
        "MÊS",
        "ANO",
        "SITUAÇÃO DO PEDIDO",
    ]

    ws.append(cabecalhos)

    for item in resultados:
        ws.append(
            [
                item["RE"],
                item["SIGLA"],
                "",                   # TIPO
                "",                   # CTE
                "",                   # VINCULAR/ ACERTO
                item["ORDEM"],
                "AUTORIZADO",
                "",                   # DT DE FINALIZAÇÃO
                "",                   # DIAS FALTANTES
                "",                   # SITUAÇÃO DE COLETA
                item["UNIDADE"],
                item["EMAIL"],        # E-MAIL
                "",                   # MÊS
                "",                   # ANO
                "",                   # SITUAÇÃO DO PEDIDO
            ]
        )

    azul_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="8EDDE1",
    )
    verde_autorizado = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )
    fonte_autorizado = Font(
        color="006100",
    )
    borda_fina = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for celula in ws[1]:
        celula.fill = azul_cabecalho
        celula.font = Font(bold=True)
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        celula.border = borda_fina

    for linha in range(2, ws.max_row + 1):
        for coluna in range(1, 16):
            celula = ws.cell(linha, coluna)
            celula.border = borda_fina
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        situacao = ws.cell(linha, 7)
        situacao.fill = verde_autorizado
        situacao.font = fonte_autorizado

    larguras = {
        "A": 14,   # RE
        "B": 18,   # SIGLA
        "C": 14,   # TIPO
        "D": 14,   # CTE
        "E": 24,   # VINCULAR/ ACERTO
        "F": 18,   # ORDEM
        "G": 18,   # SITUAÇÃO
        "H": 22,   # DT DE FINALIZAÇÃO
        "I": 18,   # DIAS FALTANTES
        "J": 24,   # SITUAÇÃO DE COLETA
        "K": 38,   # UNIDADE
        "L": 28,   # E-MAIL
        "M": 14,   # MÊS
        "N": 12,   # ANO
        "O": 24,   # SITUAÇÃO DO PEDIDO
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{ws.max_row}"

    wb.save(caminho)

    return caminho


def recuperar_home(
    page: Page,
    usuario: str,
    senha: str,
    log: Callable[[str], None],
) -> None:
    log("Recuperando a página para a próxima tentativa...")

    realizar_login(
        page,
        usuario,
        senha,
        log,
    )


def executar_automacao(
    execucoes: list[dict],
    headless: bool,
    modo_teste: bool,
    continuar_em_erro: bool,
    log: Callable[[str], None],
) -> dict:
    erros_fixos = validar_arquivos_fixos()

    emails_unidades = carregar_emails_unidades()

    if erros_fixos:
        raise ValueError("\n".join(erros_fixos))

    usuario, senha = carregar_login()

    PASTA_PERFIL.mkdir(
        parents=True,
        exist_ok=True,
    )

    autorizadas: list[dict] = []
    detalhes: list[dict] = []
    sucessos = 0
    falhas = 0

    with sync_playwright() as p:
        contexto = p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_PERFIL),
            headless=headless,
            slow_mo=120 if not headless else 0,
            permissions=[
                "clipboard-read",
                "clipboard-write",
            ],
        )

        page = (
            contexto.pages[0]
            if contexto.pages
            else contexto.new_page()
        )

        page.set_default_timeout(TIMEOUT)

        try:
            realizar_login(
                page,
                usuario,
                senha,
                log,
            )

            total = len(execucoes)

            for posicao, item in enumerate(
                execucoes,
                start=1,
            ):
                sigla = item["sigla"]

                log("")
                log(
                    f"[{posicao}/{total}] {sigla} "
                    f"({item['sequencia']}/{item['total_sigla']})"
                )

                sucesso_item = False
                ultimo_erro = None
                ultima_mensagem = ""

                for tentativa in range(
                    1,
                    MAX_TENTATIVAS + 1,
                ):
                    geracao_iniciada = False

                    log(
                        f"Tentativa {tentativa}/"
                        f"{MAX_TENTATIVAS}"
                    )

                    try:
                        # -----------------------------------------
                        # CAMINHO COMPLETO
                        # -----------------------------------------

                        abrir_solicitacao_coleta(
                            page,
                            log,
                        )

                        log(
                            f"CNPJ do remetente: "
                            f"{item['cnpj_remetente']}"
                        )

                        preencher_formulario(
                            page,
                            item["cnpj_remetente"],
                        )

                        # -----------------------------------------
                        # UNIDADE COLETORA
                        # -----------------------------------------

                        log(
                            "Capturando Unidade coletora..."
                        )

                        unidade_coletora = (
                            capturar_unidade_coletora(
                                page
                            )
                        )

                        log(
                            f"Unidade coletora: "
                            f"{unidade_coletora}"
                        )

                        # -----------------------------------------
                        # E-MAIL DA UNIDADE
                        # -----------------------------------------

                        email_unidade = (
                            emails_unidades.get(
                                unidade_coletora,
                                "",
                            )
                        )

                        if email_unidade:
                            log(
                                f"E-mail da unidade: "
                                f"{email_unidade}"
                            )
                        else:
                            log(
                                "AVISO: E-mail não encontrado "
                                f"para {unidade_coletora}"
                            )

                        # -----------------------------------------
                        # MODO TESTE
                        # -----------------------------------------

                        if modo_teste:
                            sucessos += 1

                            detalhes.append(
                                {
                                    "SIGLA": sigla,
                                    "ORDEM": "",
                                    "SITUAÇÃO": (
                                        "TESTE CONCLUÍDO"
                                    ),
                                    "MENSAGEM": (
                                        "Formulário preenchido "
                                        "sem gerar coleta."
                                    ),
                                }
                            )

                            log(
                                "Modo teste: "
                                "formulário preenchido."
                            )

                            sucesso_item = True
                            break

                        # -----------------------------------------
                        # GERAÇÃO DA ORDEM
                        # -----------------------------------------
                        #
                        # A partir daqui NÃO fazemos nova
                        # tentativa automática caso aconteça erro,
                        # pois o clique pode ter gerado a coleta.
                        # -----------------------------------------

                        geracao_iniciada = True

                        numero_ordem = (
                            gerar_coleta_e_capturar_ordem(
                                page
                            )
                        )

                        data_geracao = (
                            datetime.now().strftime(
                                "%d/%m/%Y"
                            )
                        )

                        autorizadas.append(
                            {
                                "RE": data_geracao,
                                "SIGLA": sigla,
                                "ORDEM": numero_ordem,
                                "UNIDADE": (
                                    unidade_coletora
                                ),
                                "EMAIL": email_unidade,
                            }
                        )

                        detalhes.append(
                            {
                                "SIGLA": sigla,
                                "ORDEM": numero_ordem,
                                "UNIDADE": (
                                    unidade_coletora
                                ),
                                "SITUAÇÃO": "AUTORIZADO",
                                "MENSAGEM": (
                                    "Coleta gerada "
                                    "com sucesso."
                                ),
                            }
                        )

                        sucessos += 1
                        sucesso_item = True

                        log(
                            f"Ordem gerada: "
                            f"{numero_ordem}"
                        )

                        break

                    except Exception as erro:
                        ultimo_erro = erro

                        ultima_mensagem = (
                            f"{type(erro).__name__}: "
                            f"{erro}"
                        )

                        log(
                            f"Falha na tentativa "
                            f"{tentativa}/"
                            f"{MAX_TENTATIVAS}: "
                            f"{ultima_mensagem}"
                        )

                        # -----------------------------------------
                        # NÃO RETENTA SE JÁ INICIOU A GERAÇÃO
                        # -----------------------------------------

                        if geracao_iniciada:
                            log(
                                "A falha ocorreu após iniciar "
                                "a geração da coleta."
                            )

                            log(
                                "A ordem NÃO será repetida "
                                "automaticamente para evitar "
                                "duplicidade."
                            )

                            break

                        # -----------------------------------------
                        # AINDA TEM TENTATIVA?
                        # -----------------------------------------

                        if (
                            tentativa
                            < MAX_TENTATIVAS
                        ):
                            log(
                                "Preparando nova tentativa..."
                            )

                            try:
                                recuperar_home(
                                    page,
                                    usuario,
                                    senha,
                                    log,
                                )

                            except Exception as erro_rec:
                                ultima_mensagem = (
                                    "Falha ao recuperar a "
                                    "página: "
                                    f"{type(erro_rec).__name__}: "
                                    f"{erro_rec}"
                                )

                                log(
                                    ultima_mensagem
                                )

                                ultimo_erro = erro_rec

                                break

                            continue

                        log(
                            "Número máximo de tentativas "
                            "atingido."
                        )

                # =============================================
                # SÓ REGISTRA FALHA DEPOIS DAS TENTATIVAS
                # =============================================

                if not sucesso_item:
                    falhas += 1

                    detalhes.append(
                        {
                            "SIGLA": sigla,
                            "ORDEM": "",
                            "SITUAÇÃO": "FALHA",
                            "MENSAGEM": (
                                ultima_mensagem
                            ),
                        }
                    )

                    log(
                        f"Ordem da sigla {sigla} "
                        "marcada como FALHA."
                    )

                    if not continuar_em_erro:
                        raise RuntimeError(
                            ultima_mensagem
                        ) from ultimo_erro

                    # Deixa a página pronta para
                    # a próxima ordem.
                    try:
                        recuperar_home(
                            page,
                            usuario,
                            senha,
                            log,
                        )

                    except Exception:
                        log(
                            "Não foi possível recuperar "
                            "a página após a falha final."
                        )

        finally:
            contexto.close()

    arquivo_resultado = None
    arquivo_erros = None

    if autorizadas:
        arquivo_resultado = criar_planilha_resultados(
            autorizadas
        )

    arquivo_erros = criar_planilha_erros(
        detalhes
    )

    return {
        "total": len(execucoes),
        "sucessos": sucessos,
        "falhas": falhas,
        "detalhes": detalhes,

        "arquivo_resultado": (
            str(arquivo_resultado)
            if arquivo_resultado
            else None
        ),

        "arquivo_erros": (
            str(arquivo_erros)
            if arquivo_erros
            else None
        ),
    }
